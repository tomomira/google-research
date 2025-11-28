"""Excel形式でデータを出力するモジュール

このモジュールは、抽出されたデータをExcel形式で出力する機能を提供します。
openpyxlを使用して、フォーマット適用や列幅の自動調整を行います。
"""

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import Settings
from config.constants import ERROR_MESSAGES, SUCCESS_MESSAGES
from output.formatter import OutputData
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelWriter:
    """Excel形式でデータを出力するクラス

    pandasとopenpyxlを使用して、データをExcel形式で出力します。
    フォーマット適用や列幅の自動調整を行います。
    """

    def __init__(self):
        """初期化

        ExcelWriterインスタンスを初期化します。
        """
        self.output_dir = Settings.OUTPUT_DIR
        logger.info("ExcelWriter initialized")

    def write(
        self,
        data_list: list[OutputData],
        filename: str,
        apply_format: bool = True
    ) -> Optional[Path]:
        """データをExcelに書き込み

        Args:
            data_list: 出力データのリスト
            filename: 出力ファイル名（拡張子含む）
            apply_format: フォーマットを適用するかどうか（デフォルト: True）

        Returns:
            出力したファイルのパス。失敗した場合はNone

        Raises:
            ValueError: データリストが空の場合
        """
        if not data_list:
            logger.error("Data list is empty")
            raise ValueError(ERROR_MESSAGES["empty_data"])

        logger.info(f"Writing {len(data_list)} items to Excel: {filename}")

        # 出力パスの構築
        output_path = Settings.get_output_path(filename)

        try:
            # DataFrameに変換
            df = pd.DataFrame([data.to_dict() for data in data_list])

            # 列名の日本語化
            column_names = {
                "rank": "順位",
                "title": "タイトル",
                "url": "URL",
                "description": "説明文",
                "phone": "電話番号",
                "email": "メールアドレス",
                "postal_code": "郵便番号",
                "prefecture": "都道府県",
                "fax": "FAX",
                "company_name": "会社名・店舗名",
                "sns_twitter": "Twitter",
                "sns_facebook": "Facebook",
                "sns_instagram": "Instagram"
            }
            df = df.rename(columns=column_names)

            # Excelに書き込み
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='検索結果', index=False)

                # フォーマット適用
                if apply_format:
                    workbook = writer.book
                    worksheet = writer.sheets['検索結果']
                    self._apply_format(worksheet, len(df))
                    self._auto_adjust_column_width(worksheet)

            logger.info(f"Successfully wrote to Excel: {output_path}")
            logger.info(SUCCESS_MESSAGES["excel_saved"].format(path=output_path))

            return output_path

        except Exception as e:
            logger.error(f"Failed to write Excel file: {e}", exc_info=True)
            return None

    def _apply_format(self, worksheet, data_rows: int) -> None:
        """セルのフォーマットを適用

        Args:
            worksheet: openpyxlのワークシート
            data_rows: データ行数
        """
        logger.debug("Applying cell format")

        # ヘッダー行のスタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # データ行のスタイル
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 罫線のスタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ヘッダー行（1行目）のフォーマット
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # データ行のフォーマット
        for row in range(2, data_rows + 2):  # ヘッダー行の次から
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border

        # 行の高さを設定
        worksheet.row_dimensions[1].height = 25  # ヘッダー行
        for row in range(2, data_rows + 2):
            worksheet.row_dimensions[row].height = 40  # データ行

        logger.debug("Cell format applied")

    def _auto_adjust_column_width(self, worksheet) -> None:
        """列幅を自動調整

        Args:
            worksheet: openpyxlのワークシート
        """
        logger.debug("Auto-adjusting column widths")

        # 列ごとの推奨幅を設定
        column_widths = {
            "順位": 8,
            "タイトル": 40,
            "URL": 50,
            "説明文": 50,
            "電話番号": 20,
            "メールアドレス": 30,
            "郵便番号": 12,
            "都道府県": 12,
            "FAX": 20,
            "会社名・店舗名": 30,
            "Twitter": 40,
            "Facebook": 40,
            "Instagram": 40
        }

        for col in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col)
            header_cell = worksheet.cell(row=1, column=col)
            header_value = header_cell.value

            # 推奨幅を設定
            if header_value in column_widths:
                width = column_widths[header_value]
            else:
                # デフォルト幅
                width = 15

            worksheet.column_dimensions[column_letter].width = width

        logger.debug("Column widths adjusted")

    def create_summary_sheet(
        self,
        workbook: Workbook,
        total_results: int,
        keyword: str,
        timestamp: str
    ) -> None:
        """サマリーシートを作成

        検索情報の概要を記載したシートを作成します。

        Args:
            workbook: openpyxlのワークブック
            total_results: 検索結果の総数
            keyword: 検索キーワード
            timestamp: タイムスタンプ
        """
        logger.debug("Creating summary sheet")

        # サマリーシートの作成
        summary_sheet = workbook.create_sheet(title="サマリー", index=0)

        # タイトル
        summary_sheet['A1'] = 'Google検索リサーチ結果'
        summary_sheet['A1'].font = Font(bold=True, size=14)

        # 検索情報
        summary_sheet['A3'] = '検索キーワード:'
        summary_sheet['B3'] = keyword

        summary_sheet['A4'] = '検索日時:'
        summary_sheet['B4'] = timestamp

        summary_sheet['A5'] = '検索結果件数:'
        summary_sheet['B5'] = total_results

        # 列幅調整
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 50

        logger.debug("Summary sheet created")
